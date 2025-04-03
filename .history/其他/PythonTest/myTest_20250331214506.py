# 导入模块
import openpyxl

# 1、打开文件，生成文件对象
workbook = openpyxl.load_workbook("其他\Test\excel\演员单位.ts.xlsx")
# 2、获取活动表对象
sheet = workbook.active
print(f'当前活动表为：{sheet}')

# 3、遍历
# for i in range(3, sheet.max_row):
#     for j in range(1, sheet.max_column):
#         value = sheet.cell(row=i, column=j).value
#         if value is not None:
#             print(value)

for row_idx, row in enumerate(sheet.iter_rows(min_row=3), 3):
    for col_idx, cell in enumerate(row, 1):
        value = cell.value
        if value is not None:
            print(f'{row_idx},{col_idx},{value}')
