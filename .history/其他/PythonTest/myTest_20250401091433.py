# 导入模块
import openpyxl

# 1、打开文件，生成文件对象
workbook = openpyxl.load_workbook("其他\Test\excel\演员单位.ts.xlsx", data_only=True)
# 2、获取活动表对象
sheet = workbook.active
print(f'当前活动表为：{sheet}')

cell = sheet.cell(3, 2)  # 得到A1单元格对象
# 4、获取字体样式对象
font = cell.font  # 得到该单元格的字体样式对象
color = font.color.rgb  # 得到该单元格的字体颜色对象
# 5、打印该单元格数据的字体样式
print('下面是该单元格数据的字体样式：')
print(color)  # 字体颜色

# 3、遍历
# for i in range(3, sheet.max_row):
#     for j in range(1, sheet.max_column):
#         value = sheet.cell(i, j).value
#         if value is not None:
#             print(value)

# print(
#     "*******************************************************************************************************"
# )
# for row_idx, row in enumerate(sheet.iter_rows(min_row=3), 3):
#     for col_idx, cell in enumerate(row, 1):
#         value = cell.value
#         if value is not None:
#             print(f'{row_idx},{col_idx},{value}')

# for row_idx, row in enumerate(sheet.rows, 1):
#     if row_idx <= 2: continue
#     for col_idx, cell in enumerate(row, 1):
#         value = cell.value
#         if value is not None:
#             print(f'{row_idx},{col_idx},{value}')
