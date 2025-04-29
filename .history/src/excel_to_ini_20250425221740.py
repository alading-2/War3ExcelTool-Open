# excel_to_ini.py
# 实现 Excel 批量读取与物编表格识别
# 仅实现本阶段功能，后续功能待用户确认后再开发

import os
import traceback
from typing import List, Dict, Any
from openpyxl import load_workbook

# 支持的物编类型 sheet 名
OBJECT_TYPE_MAPPING = {
    'unit': ['单位', 'unit'],
    'item': ['物品', 'item'],
    'ability': ['技能', 'ability']
}


def is_baize_table(file_name: str, sheet_name: str) -> bool:
    """
    判断是否为白泽 table 文件的 unit/ability/item sheet
    """
    if file_name.lower() == 'table' and any(
            sheet_name.lower() == t for v in OBJECT_TYPE_MAPPING.values()
            for t in v):
        return True
    return False


def is_common_wubian_sheet(sheet_keys: List[str]) -> bool:
    """
    判断 sheet 的第二行 key 是否包含 id 和 _parent
    """
    keys_lower = [k.strip().lower() for k in sheet_keys]
    return 'id' in keys_lower and '_parent' in keys_lower


def find_excel_files(input_dir: str) -> List[str]:
    """
    递归查找目录下所有 .xlsx 文件
    """
    excel_files = []
    for root, _, files in os.walk(input_dir):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~$'):
                excel_files.append(os.path.join(root, file))
    return excel_files


def parse_excel_for_wubian_tables(excel_path: str) -> List[Dict[str, Any]]:
    """
    解析单个 Excel 文件，识别所有需转化的 sheet
    返回: List[Dict]，每项包含 file, sheet, keys
    """
    result = []
    file_name = os.path.splitext(os.path.basename(excel_path))[0]
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                # 读取前两行
                rows = list(
                    ws.iter_rows(min_row=1, max_row=2, values_only=True))
                if len(rows) < 2:
                    continue
                keys = [
                    str(k).strip() if k is not None else '' for k in rows[1]
                ]
                # 白泽 table 文件
                if is_baize_table(file_name, sheet_name):
                    result.append({
                        'file': excel_path,
                        'sheet': sheet_name,
                        'keys': keys
                    })
                    continue
                # 通用物编表格
                if is_common_wubian_sheet(keys):
                    result.append({
                        'file': excel_path,
                        'sheet': sheet_name,
                        'keys': keys
                    })
            except Exception as e:
                print(
                    f"[错误] 解析 sheet 时异常: 文件={excel_path}, sheet={sheet_name}")
                traceback.print_exc()
    except Exception as e:
        print(f"[错误] 解析 Excel 文件异常: 文件={excel_path}")
        traceback.print_exc()
    return result


def batch_parse_excels(input_dir: str) -> List[Dict[str, Any]]:
    """
    批量解析目录下所有 Excel，返回所有需转化的表格信息
    """
    all_results = []
    excel_files = find_excel_files(input_dir)
    for excel_path in excel_files:
        res = parse_excel_for_wubian_tables(excel_path)
        all_results.extend(res)
    return all_results


def main():
    import argparse
    parser = argparse.ArgumentParser(description="批量识别需转化的物编表格")
    parser.add_argument('-i', '--input', required=True, help='输入 Excel 文件夹路径')
    args = parser.parse_args()
    results = batch_parse_excels(args.input)
    print("识别到以下需转化的表格：")
    for item in results:
        print(
            f"文件: {item['file']} | sheet: {item['sheet']} | keys: {item['keys']}"
        )


if __name__ == '__main__':
    main()
