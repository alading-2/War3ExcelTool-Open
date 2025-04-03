import re  # 导入正则表达式模块，用于字符串模式匹配
from ast import literal_eval
import os  # 导入操作系统模块，用于处理文件路径
import pandas as pd  # 导入pandas库，简称pd，用于数据分析和处理Excel文件
import openpyxl  # 导入openpyxl库，用于读取Excel文件的样式信息（如颜色）
from typing import Dict, List, Any, Tuple, Optional, Union  # 导入类型提示功能
from openpyxl.styles import PatternFill  # 导入Excel单元格填充样式相关功能
import logging  # 导入日志模块
from src.utils.logger import get_logger  # 从项目的utils目录导入自定义日志获取函数

# 获取当前模块的日志记录器
logger = get_logger(__name__)  # __name__是Python的特殊变量，表示当前模块名称


class ExcelParser:
    """解析Excel文件并提取数据的类
    
    这个类负责读取Excel文件的内容，提取其中的数据、结构和颜色信息，
    为后续生成TypeScript代码做准备。
    """

    def __init__(self):
        """初始化ExcelParser类的新实例
        
        创建一个日志记录器，用于记录Excel解析过程中的信息
        """
        # 获取一个与当前类关联的日志记录器
        self.logger = logging.getLogger(__name__)

    def parse_excel(
            self,
            file_path: str) -> List[Tuple[str, Dict[str, Any], List[Dict]]]:
        """
        解析Excel文件，提取数据和结构信息
        
        这个方法是整个解析过程的主入口，它会读取Excel文件并提取多种信息：
        1. sheetnames
        2. 元数据（如注释、键名等）
        3. 实际数据行的列表
        
        Args:
            file_path: Excel文件路径，表示要解析的Excel文件的位置
            sheet_name: 可选的工作表名称，如果不提供则解析所有工作表
            
        Returns:
            1. sheetnames
            2. 元数据（如注释、键名等）
            3. 实际数据行的列表
            
        Raises:
            FileNotFoundError: 当指定的Excel文件不存在时
            Exception: 当解析过程中发生其他错误时
        """
        # 检查文件是否存在，不存在则记录错误并抛出异常
        if not os.path.exists(file_path):
            self.logger.error(f"Excel文件不存在: {file_path}")
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")

        try:
            # 获取Excel文件的所有sheet名称
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names

            results = []

            # 遍历所有sheet并解析
            for sheet_name in sheet_names:
                self.logger.info(f"解析工作表: {sheet_name}")

                # 使用pandas读取指定sheet的内容
                df = pd.read_excel(file_path,
                                   sheet_name=sheet_name,
                                   header=None)
                # 提取第一行作为注释信息（列说明），首尾去除空格，不是空行
                comments = [
                    str(val).strip() for val in df.iloc[0].tolist()
                    if pd.isna(val) == False and val != ""
                ]

                # 提取第二行作为数据的键名，首尾去除空格
                keys = [
                    str(val).strip() for val in df.iloc[1].tolist()
                    if pd.isna(val) == False and val != ""
                ]

                # 提取第三行数据推断类型
                types = self.infer_types(df)

                #TODO 预处理
                # # 预处理指令列表
                # preprocessor_instruction = ['#default', '#path']
                # # 存储预处理指令的字典
                # preprocessors = {}
                # # 搜索前3行是否有预处理指令
                # for idx, row in df.iloc[0:2].iterrows():
                #     for i, val in enumerate(row):
                #         # 搜索所有预处理指令 比如：#default 1000
                #         prepval = re.findall(r'#\w+\W?(\w+)?', val)
                #         if prepval:
                #             prep = re.search(r'#\w+\W?', prepval) # 提取预处理指令
                #             prep_val = re.search(r'\W(\w+)?', prepval) # 提取预处理指令值
                #             if prep in preprocessor_instruction:
                #                 preprocessors.keys[i].prep = prep
                #                 preprocessors.keys[i].prep_val = prep_val

                # 创建元数据字典
                metadata = {
                    # "preprocessors": preprocessors,  # 预处理指令
                    "comments": comments,  # 注释信息
                    "keys": keys,  # 键名列表
                    "types": types,  # 数据类型列表
                    "filename": os.path.basename(file_path),  # 文件名
                    "sheet_name": sheet_name  # 添加sheet名称到元数据
                }

                # 解析实际数据行（第三行及以下）
                data_rows = self._parse_data_rows(df, metadata)

                # 将当前sheet的解析结果添加到列表
                results.append((sheet_name, metadata, data_rows))

            return results

        except Exception as e:
            self.logger.error(f"解析Excel文件的工作表时出错: {str(e)}")
            raise

    def infer_types(self, df: pd.DataFrame) -> List[str]:
        """
        根据一列数据推断TypeScript类型
        
        Args:
            values: 要推断类型的值列表，可以是任何Python数据类型
            
        Returns:
            返回sheet的数据类型列表
        """
        types = []
        # 遍历所有列（不包括code列），推断每列的数据类型
        for j, column in df.items():
            if pd.isnan(column[1].values()):
                continue  # 跳过空列
            # 每列的数值从第3行开始，跳过第一行和第二行
            values = column[2:].tolist()

            # 确定一列的数据类型可能的取值，使用集合自动去重
            column_types = set()
            for i, value in enumerate(values):
                if value is not None:  # 跳过None值
                    # 使用类型处理器推断值的TypeScript类型
                    type = self.infer_type_single(value)
                    if type is not None:
                        column_types.add(type)

            # 根据收集到的类型确定最终字段类型
            if len(column_types) == 0:  # 如果没有收集到类型（可能全是None）
                column_final_type = "any"  # 使用any类型
            elif len(column_types) == 1:  # 如果只有一种类型
                # iter(types):集合转换迭代器，next() 函数从迭代器中提取第一个可用元素。由于此时len(types) == 1的判断已确保集合中只有唯一元素，因此无论集合实际存储顺序如何，总能正确获取到该唯一类型。
                column_final_type = next(iter(column_types))
            else:
                # 多种类型情况，使用联合类型
                column_final_type = "any"
            # 将每列最终类型添加到列表中
            types.append(column_final_type)

        # 返回所有列的最终类型列表
        return types

    def infer_type_single(self, value: Any) -> str:
        """
        推断单个数据的TypeScript类型
        支持推断布尔值、数字、字符串、数组、对象等多种类型。
        
        Args:
            value: 要推断类型的值，可以是任何Python数据类型
            
        Returns:
            推断出的TypeScript类型名称，例如"string"、"number"、"boolean、any"等
            
        示例:
            >>> handler = TypeHandler()
            >>> handler.infer_type(42)
            "number"
            >>> handler.infer_type("Hello")
            "string"
        """

        # 如果单元格什么都不填，值是NaN
        if pd.isna(value):
            return None

        # 如果值是布尔类型，返回"boolean"
        if isinstance(value, bool):  # isinstance检查值是否为指定类型的实例
            return "boolean"

        # 如果值是整数或浮点数，返回"number"
        if isinstance(value, (int, float)):  # 可以同时检查多种类型
            return "number"

        # 如果值为""，返回"string"类型
        if value == "":
            return "string"

        # 如果值是字符串，进一步分析它可能代表的类型
        if isinstance(value, str):
            # 检查字符串是否代表数字（使用正则表达式匹配）
            # ^表示开始，-?表示可选的负号，\d+表示一个或多个数字
            # (\.\d+)?表示可选的小数部分，$表示结束
            if re.match(r'^-?\d+(\.\d+)?([eE][+-]?\d+)?$', value):
                return "number"

            # 检查字符串是否代表布尔值（"true"或"false"）
            if value.lower() in ('true', 'false'):
                return "boolean"

            # 检查字符串是否像JSON对象或数组
            if (value.startswith('{') and value.endswith('}')) or \
                (value.startswith('[') and value.endswith(']')):
                return "any"

            # 其他字符串一律视为string类型
            return "string"

        # 默认情况下返回"any"类型
        return "any"

    def _parse_data_rows(self, df: pd.DataFrame,
                         metadata: Dict[str, Any]) -> List[Dict]:
        """
        解析数据行，将DataFrame中的数据转换为字典列表
        
        Args:
            df: DataFrame数据，pandas的数据表格对象
            metadata: 元数据字典，包含注释、键名列表和文件名等信息
            
        Returns:
            数据行列表，每个元素是一个字典，表示一行数据
        """
        # 创建空列表，用于存储解析后的数据行
        data_rows = []

        # 遍历DataFrame中的行，从第三行开始（索引2）
        # idx是行索引，row是行数据
        for idx, row in df.iloc[2:].iterrows():
            # 检查是否为注释行（首列数据为空或者以"//"开头的行视为注释，将被跳过）
            # str(): 用于将对象转换为字符串。strip(): 去除字符串两端的空格
            # startswith(): 检查字符串是否以指定的前缀开头
            if row.iloc[0] == "" or row.iloc[0].strip().startswith("//"):
                continue  # 跳过本次循环，处理下一行

            # 创建行数据字典
            row_data = {}

            # i是列索引，value是单元格的值
            for i, value in enumerate(row):
                # print(value, type(value))
                # 跳过空数据单元格/nan，
                if pd.isna(value) or value == "":
                    continue

                # 检查列索引是否在键列表范围内
                keys = metadata.get("keys")
                types = metadata.get("types")
                if i < len(keys):
                    # 获取当前列的键名
                    key = keys[i]
                    # 处理单元格的值
                    value = self._value_handle(value, str(types[i]))
                    # 将值添加到行数据字典中
                    row_data[key] = value

            # 将处理好的行数据添加到结果列表中
            data_rows.append(row_data)

        # 返回所有数据行的列表
        return data_rows

    def _value_handle(self, value: Any, type: str) -> Any:
        """
        处理Excel单元格的值
        
        根据单元格的类型（数字、字符串、对象、数组等）进行处理，
        返回适当的格式化值。
        
        Args:
            value: 单元格的值，可以是多种类型（数字、字符串、对象、数组等）
            
        Returns:
            处理后的值，适用于TypeScript代码生成
        """
        # 处理数字类型值
        if type in (int, float):
            return value

        if type == "boolean":
            return str(value).lower()

        # 处理字符串类型值
        if type in ("string", "any"):
            # 先去掉首尾空白，再去掉多余的双引号
            if value != "":
                value = value.strip().strip('"')

            # # 检查字符串是否代表数字（使用正则表达式匹配）
            # # ^表示开始，-?表示可选的负号，\d+表示一个或多个数字
            # # (\.\d+)?表示可选的小数部分，$表示结束
            # if re.fullmatch(r'^-?\d+(\.\d+)?([eE][+-]?\d+)?$', value):
            #     return f"{value}"

            # # 检查字符串是否代表布尔值（"true"或"false"）
            # if value.lower() in ('true', 'false'):
            #     return f"{value.lower()}"

            # 如果值是对象/数组的字符串，用正则表达式获取对象/数组
            if (value.startswith('{') and value.endswith('}')) or \
                (value.startswith('[') and value.endswith(']')):
                # 步骤 1：处理键的引号问题（给键添加双引号）
                # 结果：'{"strength_p":3, "attack_p":5,}'
                value = re.sub(r'(\w+):', r'"\1":', value)

                # 步骤 2：处理末尾的逗号问题
                # 结果：'{"strength_p":3, "attack_p":5}'
                value = re.sub(r',\s*}', '}', value)

                # 步骤 3：安全解析为字典
                value = literal_eval(value)
                return value
            # 普通字符串直接返回
            return value

        # 其他类型直接返回原始值
        return value

    def _extract_colors(self, file_path: str,
                        sheet_name: str) -> Dict[str, str]:
        """
        提取Excel文件中指定工作表的单元格颜色信息
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            颜色映射字典，键为单元格位置标识（如'A1'），值为颜色代码（如'FF0000'表示红色）
        """
        try:
            # 使用openpyxl加载工作簿
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            # 获取指定的工作表
            sheet = workbook[sheet_name]

            # 创建颜色映射字典
            color_map = {}

            # 遍历工作表中的所有单元格
            for row in sheet.iter_rows():
                for cell in row:
                    # 检查单元格是否有填充色
                    if cell.fill and cell.fill.fill_type == 'solid':
                        # 获取填充色的RGB值
                        fill_color = cell.fill.start_color.rgb
                        # 如果有有效的颜色值（不是默认值00000000）
                        if fill_color and fill_color != '00000000':
                            # 将单元格位置和颜色添加到映射中
                            color_map[cell.coordinate] = fill_color

            return color_map
        except Exception as e:
            self.logger.error(f"提取单元格颜色时出错: {str(e)}")
            # 出错时返回空字典
            return {}
