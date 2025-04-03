import re  # 导入正则表达式模块，用于字符串模式匹配
from ast import literal_eval
import os  # 导入操作系统模块，用于处理文件路径
import pandas as pd  # 导入pandas库，简称pd，用于数据分析和处理Excel文件
import openpyxl  # 导入openpyxl库，用于读取Excel文件的样式信息（如颜色）
from typing import Dict, List, Any, Tuple, Optional, Union  # 导入类型提示功能
from openpyxl.styles import PatternFill  # 导入Excel单元格填充样式相关功能
import logging  # 导入日志模块
from src.utils.logger import get_logger  # 从项目的utils目录导入自定义日志获取函数

# 获取当前模块的日志记录器
logger = get_logger(__name__)  # __name__是Python的特殊变量，表示当前模块名称


class ExcelParser:
    """解析Excel文件并提取数据的类
    
    这个类负责读取Excel文件的内容，提取其中的数据、结构和颜色信息，
    为后续生成TypeScript代码做准备。
    """

    def __init__(self):
        """初始化ExcelParser类的新实例
        
        创建一个日志记录器，用于记录Excel解析过程中的信息
        """
        # 获取一个与当前类关联的日志记录器
        self.logger = logging.getLogger(__name__)

    def parse_excel(
        self, file_path: str
    ) -> List[Tuple[str, Dict[str, Any], List[str], List[Dict], Dict[str,
                                                                     str]]]:
        """
        解析Excel文件，提取数据和结构信息
        
        这个方法是整个解析过程的主入口，它会读取Excel文件并提取多种信息：
        1. 元数据（如注释、键名等）
        2. 数据字段的键名列表
        3. 实际数据行的列表
        4. 单元格颜色信息
        
        Args:
            file_path: Excel文件路径，表示要解析的Excel文件的位置
            sheet_name: 可选的工作表名称，如果不提供则解析所有工作表
            
        Returns:
            如果指定了sheet_name，返回一个包含四个元素的元组(Tuple)：
            - 元数据: 字典，包含注释、键名列表和文件名等信息
            - 键列表: 列表，数据字段的键名列表（不包括ID列）
            - 数据列表: 列表，每个元素是一个字典，表示一行数据
            - 颜色映射: 字典，单元格位置到颜色代码的映射
            
            如果没有指定sheet_name，返回一个列表，每个元素包含五个元素的元组(Tuple)：
            - sheet名称: 字符串，工作表的名称
            - 元数据: 字典，包含注释、键名列表和文件名等信息
            - 键列表: 列表，数据字段的键名列表（不包括ID列）
            - 数据列表: 列表，每个元素是一个字典，表示一行数据
            - 颜色映射: 字典，单元格位置到颜色代码的映射
            
        Raises:
            FileNotFoundError: 当指定的Excel文件不存在时
            Exception: 当解析过程中发生其他错误时
        """
        # 检查文件是否存在，不存在则记录错误并抛出异常
        if not os.path.exists(file_path):
            self.logger.error(f"Excel文件不存在: {file_path}")
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")

        try:
            # 获取Excel文件的所有sheet名称
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names

            results = []

            # 遍历所有sheet并解析
            for sheet_name in sheet_names:
                self.logger.info(f"解析工作表: {sheet_name}")

                # 使用pandas读取指定sheet的内容
                df = pd.read_excel(file_path,
                                   sheet_name=sheet_name,
                                   header=None)

                # 提取第一行作为注释信息（列说明）
                comments = df.iloc[0].tolist()

                # 提取第二行作为数据的键名
                keys = df.iloc[1].tolist()

                # 解析实际数据行（第三行及以下）
                data_rows = self._parse_data_rows(df, keys)

                # 读取当前sheet的单元格颜色信息
                color_map = self._extract_colors(file_path, sheet_name)

                # 创建元数据字典
                metadata = {
                    "comments": comments,
                    "keys": keys,
                    "filename": os.path.basename(file_path),
                    "sheet_name": sheet_name  # 添加sheet名称到元数据
                }

                # 将当前sheet的解析结果添加到列表
                results.append(
                    (sheet_name, metadata, keys[1:], data_rows, color_map))

            return results

        except Exception as e:
            self.logger.error(f"解析Excel文件的工作表时出错: {str(e)}")
            raise

    def _parse_data_rows(self, df: pd.DataFrame, keys: List) -> List[Dict]:
        """
        解析数据行，将DataFrame中的数据转换为字典列表
        
        Args:
            df: DataFrame数据，pandas的数据表格对象
            keys: 键列表，表示每列的字段名
            
        Returns:
            数据行列表，每个元素是一个字典，表示一行数据
        """
        # 创建空列表，用于存储解析后的数据行
        data_rows = []

        # 遍历DataFrame中的行，从第三行开始（索引2）
        # idx是行索引，row是行数据
        for idx, row in df.iloc[2:].iterrows():
            # 检查是否为注释行（首列数据为空或者以"//"开头的行视为注释，将被跳过）
            # str(): 用于将对象转换为字符串。strip(): 去除字符串两端的空格
            # startswith(): 检查字符串是否以指定的前缀开头
            if row.iloc[0] == "" or row.iloc[0].strip().startswith("//"):
                continue  # 跳过本次循环，处理下一行

            # 创建行数据字典
            row_data = {}

            # 遍历行中的每个单元格（从第二列开始）
            # i是列索引（从1开始），value是单元格的值
            for i, value in enumerate(row):
                # 跳过空数据单元格（NaN或空字符串）
                if pd.isna(value) or value == "":
                    continue

                # 检查列索引是否在键列表范围内
                if i < len(keys):
                    # 获取当前列的键名
                    key_name = keys[i]
                    # 处理单元格的值
                    value = self._value_handle(value)
                    # 将值添加到行数据字典中
                    row_data[key_name] = value

            # 将处理好的行数据添加到结果列表中
            data_rows.append(row_data)

        # 返回所有数据行的列表
        return data_rows

    def _value_handle(self, value: Any) -> Any:
        """
        处理Excel单元格的值
        
        根据单元格的类型（数字、字符串、对象、数组等）进行处理，
        返回适当的格式化值。
        
        Args:
            value: 单元格的值，可以是多种类型（数字、字符串、对象、数组等）
            
        Returns:
            处理后的值，适用于TypeScript代码生成
        """
        # 处理数字类型值
        if isinstance(value, (int, float)):
            return value

        # 处理字符串类型值
        if isinstance(value, str):
            # 检查字符串是否代表数字（使用正则表达式匹配）
            # ^表示开始，-?表示可选的负号，\d+表示一个或多个数字
            # (\.\d+)?表示可选的小数部分，$表示结束
            if re.match(r'^-?\d+(\.\d+)?$', value):
                return f"{value}"

            # 检查字符串是否代表布尔值（"true"或"false"）
            if value.lower() in ('true', 'false'):
                return f"{value.lower()}"

            # 先去掉首尾空白，再去掉多余的双引号
            value = value.strip().strip('"')

            # 如果值是对象/数组的字符串，用正则表达式获取对象/数组
            if (value.startswith('{') and value.endswith('}')) or \
                (value.startswith('[') and value.endswith(']')):
                # 步骤 1：处理键的引号问题（给键添加双引号）
                # 结果：'{"strength_p":3, "attack_p":5,}'
                value = re.sub(r'(\w+):', r'"\1":', value)

                # 步骤 2：处理末尾的逗号问题
                # 结果：'{"strength_p":3, "attack_p":5}'
                value = re.sub(r',\s*}', '}', value)

                # 步骤 3：安全解析为字典
                value = literal_eval(value)
                return value
            return value

        # 其他类型直接返回原始值
        return value

    def _extract_colors(self, file_path: str,
                        sheet_name: str) -> Dict[str, str]:
        """
        提取Excel文件中指定工作表的单元格颜色信息
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            颜色映射字典，键为单元格位置标识（如'A1'），值为颜色代码（如'FF0000'表示红色）
        """
        try:
            # 使用openpyxl加载工作簿
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            # 获取指定的工作表
            sheet = workbook[sheet_name]

            # 创建颜色映射字典
            color_map = {}

            # 遍历工作表中的所有单元格
            for row in sheet.iter_rows():
                for cell in row:
                    # 检查单元格是否有填充色
                    if cell.fill and cell.fill.fill_type == 'solid':
                        # 获取填充色的RGB值
                        fill_color = cell.fill.start_color.rgb
                        # 如果有有效的颜色值（不是默认值00000000）
                        if fill_color and fill_color != '00000000':
                            # 将单元格位置和颜色添加到映射中
                            color_map[cell.coordinate] = fill_color

            return color_map
        except Exception as e:
            self.logger.error(f"提取单元格颜色时出错: {str(e)}")
            # 出错时返回空字典
            return {}
