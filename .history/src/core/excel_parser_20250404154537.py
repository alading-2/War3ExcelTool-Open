import re  # 导入正则表达式模块，用于字符串模式匹配
from ast import literal_eval
import os  # 导入操作系统模块，用于处理文件路径
import pandas as pd  # 导入pandas库，简称pd，用于数据分析和处理Excel文件
import openpyxl  # 导入openpyxl库，用于读取Excel文件的样式信息（如颜色）
from typing import Dict, List, Any, Tuple, Optional, Union  # 导入类型提示功能
from openpyxl.styles import PatternFill  # 导入Excel单元格填充样式相关功能
import logging  # 导入日志模块
from src.utils.logger import get_logger  # 从项目的utils目录导入自定义日志获取函数
from dataclasses import dataclass, field # 导入dataclass装饰器和field函数，用于创建数据类

# 获取当前模块的日志记录器
logger = get_logger(__name__)  # __name__是Python的特殊变量，表示当前模块名称

# 定义用于存储列信息的 dataclass
@dataclass # 使用dataclass装饰器自动生成特殊方法（如__init__, __repr__）
class ColumnInfo:
    """存储Excel表格中单列的元数据信息"""
    key: str                       # 列的键名 (来自第二行)
    type: str                      # 推断出的TypeScript类型
    comment: Optional[str] = None  # 列的注释 (来自第一行)，使用Optional表示可能为None
    # preprocessor: Optional[Dict[str, Any]] = None # 预留给预处理指令的空间, field(default_factory=dict) 更合适如果需要可变默认值

class ExcelParser:
    """解析Excel文件并提取数据的类
    
    这个类负责读取Excel文件的内容，提取其中的数据、结构和颜色信息，
    为后续生成TypeScript代码做准备。
    """

    def __init__(self):
        """初始化ExcelParser类的新实例
        
        创建一个日志记录器，用于记录Excel解析过程中的信息
        """
        # 获取一个与当前类关联的日志记录器
        self.logger = logging.getLogger(__name__)

    def parse_excel(
            self,
            file_path: str) -> List[Tuple[str, Dict[str, Any], List[Dict]]]:
        """
        解析Excel文件，提取数据和结构信息
        
        这个方法是整个解析过程的主入口，它会读取Excel文件并提取多种信息：
        1. 工作表名称 (sheet_name)
        2. 包含列信息和文件元数据的 metadata 字典
        3. 实际数据行的列表 (data_rows)
        
        Args:
            file_path: Excel文件路径，表示要解析的Excel文件的位置
            
        Returns:
            一个包含元组的列表，每个元组代表一个工作表，包含:
            (sheet_name, metadata, data_rows)
            其中 metadata 结构为:
            {
                "filename": str,
                "sheet_name": str,
                "columns": List[ColumnInfo] # 列信息对象列表
            }
            
        Raises:
            FileNotFoundError: 当指定的Excel文件不存在时
            Exception: 当解析过程中发生其他错误时
        """
        # 检查文件是否存在，不存在则记录错误并抛出异常
        if not os.path.exists(file_path):
            self.logger.error(f"Excel文件不存在: {file_path}")
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")

        try:
            # 获取Excel文件的所有sheet名称
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names

            results = []

            # 遍历所有sheet并解析
            for sheet_name in sheet_names:
                self.logger.info(f"解析工作表: {sheet_name}")

                # 使用pandas读取指定sheet的内容
                df = pd.read_excel(file_path,
                                   sheet_name=sheet_name,
                                   header=None)
                # 检查DataFrame是否为空或行数不足以提取键名
                if df.empty or len(df) < 2:
                    self.logger.warning(f"工作表 '{sheet_name}' 在文件 '{os.path.basename(file_path)}' 中为空或缺少必要的行（至少需要2行），跳过。")
                    continue

                # --- 重构元数据提取 ---
                columns_info: List[ColumnInfo] = [] # 用于存储所有列信息的列表
                # 提取原始行数据，处理行数不足的情况
                raw_comments = df.iloc[0].tolist() if len(df) > 0 else []
                raw_keys = df.iloc[1].tolist() # 键名行保证存在 (len(df) >= 2)
                # 推断类型，如果行数不足3，则无法推断，后续会处理
                raw_types = self.infer_types(df) # infer_types 内部会处理行数不足的情况

                key_map = {} # 内部使用，方便通过键名查找列信息

                # 遍历第二行（键名行）来确定有效的列及其信息
                for i, key_any in enumerate(raw_keys):
                    key = str(key_any).strip() # 获取键名并去除首尾空格
                    # 只处理键名不为空的列
                    if pd.notna(key_any) and key != "":
                        # 获取对应的注释（第一行），注意索引边界
                        comment_any = raw_comments[i] if i < len(raw_comments) else None
                        comment = str(comment_any).strip() if pd.notna(comment_any) and str(comment_any).strip() != "" else None

                        # 获取对应的推断类型（来自 infer_types 的结果），注意索引边界
                        col_type = raw_types[i] if i < len(raw_types) else "any" # 如果类型列表较短，则默认为any

                        # TODO: 在这里添加预处理指令的解析逻辑 (如果需要)
                        # preprocessor_info = self._parse_preprocessor(raw_comments, raw_keys, i)

                        # 创建ColumnInfo对象
                        column = ColumnInfo(
                            key=key,
                            type=col_type,
                            comment=comment
                            # preprocessor=preprocessor_info # 将解析出的预处理信息赋给ColumnInfo
                        )
                        columns_info.append(column) # 添加到列信息列表
                        # 记录键名到列信息的映射，方便后续查找
                        # 如果有重复键名，这里会覆盖，可能需要处理或发出警告
                        if key in key_map:
                            self.logger.warning(f"在工作表 '{sheet_name}' 中发现重复的键名: '{key}'。后续处理将使用最后出现的列信息。")
                        key_map[key] = column

                # --- 结束重构元数据提取 ---

                # 创建新的 metadata 结构
                metadata = {
                    "filename": os.path.basename(file_path), # 文件名
                    "sheet_name": sheet_name,                 # 工作表名
                    "columns": columns_info,                  # 包含所有ColumnInfo对象的列表
                    "_key_map": key_map                       # 内部使用的键名到列信息的映射
                }

                # 解析实际数据行（第三行及以下）
                # 注意：_parse_data_rows 现在需要使用新的 metadata 结构
                data_rows = self._parse_data_rows(df, metadata)

                # 清理掉内部使用的映射，不暴露给外部
                del metadata["_key_map"]

                # 将当前工作表的解析结果添加到最终列表中
                results.append((sheet_name, metadata, data_rows))

            return results

        except Exception as e:
            self.logger.error(f"解析Excel文件的工作表时出错: {str(e)}")
            raise

    def infer_types(self, df: pd.DataFrame) -> List[str]:
        """
        根据DataFrame数据推断每列的TypeScript类型列表。
        返回与DataFrame列数一致的类型列表。
        
        Args:
            df: 包含Excel数据的DataFrame。
            
        Returns:
            一个字符串列表，包含每列推断出的TypeScript类型。
            列表长度与df的列数相同。 如果行数不足（<3），则所有列类型为 "any"。
        """
        num_columns = df.shape[1] # 获取DataFrame的列数

        # 如果行数不足3（没有数据行），无法进行类型推断，所有列视为any
        if len(df) < 3:
            self.logger.warning("数据行不足（少于3行），无法进行类型推断，所有列将默认为 'any' 类型。")
            return ["any"] * num_columns

        types = [] # 存储每列最终推断出的类型
        # 遍历所有列的索引
        for j in range(num_columns):
            # 检查第二行（键名行）对应的列是否有键名
            key_cell = df.iloc[1, j]
            # 没有键名的列，或者键名为空字符串的列，我们不关心其类型，设为any
            if pd.isna(key_cell) or str(key_cell).strip() == "":
                 types.append("any")
                 continue

            # 获取该列从第三行开始的所有数据值，用于类型推断
            values = df.iloc[2:, j].tolist()

            column_types = set() # 存储该列遇到的所有非空值的推断类型
            has_non_na = False   # 标记该列是否有非空（非NA）值

            for value in values:
                if pd.notna(value): # 只处理非空值
                    has_non_na = True
                    # 推断单个值的类型
                    inferred_type = self.infer_type_single(value)
                    # infer_type_single 对 NA 返回 None，这里不需要再次检查
                    # 但它可能返回 "any" 或其他具体类型
                    if inferred_type is not None: # 理论上 pd.notna(value) 为真时不会是 None
                        column_types.add(inferred_type)

            # 确定该列的最终类型
            if not has_non_na or not column_types:
                # 如果该列全是空值，或者所有非空值都无法推断出类型（例如都是复杂对象）
                column_final_type = "any"
            elif len(column_types) == 1:
                # 如果只遇到一种类型
                column_final_type = next(iter(column_types))
            elif "any" in column_types:
                # 如果推断出的类型中包含 'any'，则最终类型为 'any'
                column_final_type = "any"
            elif column_types == {"number", "boolean"}:
                 # 数字和布尔混合，TypeScript 中没有直接对应类型，用 any
                 column_final_type = "any" # 或者可以考虑 'number | boolean'，但 'any' 更简单
            elif "string" in column_types:
                 # 如果包含字符串以及其他类型（非any），通常字符串是兼容性最强的
                 # 也可以选择 'any'
                 column_final_type = "string"
            else:
                # 其他多种类型混合的情况（例如 number 和 boolean），使用 'any' 最安全
                # 或者可以生成联合类型: " | ".join(sorted(list(column_types)))
                column_final_type = "any"

            types.append(column_final_type) # 将该列的最终类型添加到列表中

        return types # 返回包含所有列类型的列表

    def infer_type_single(self, value: Any) -> Optional[str]: # 返回Optional[str]因为pd.NA应返回None
        """
        推断单个数据的TypeScript类型。
        现在对数字字符串的判断更严格，并尝试解析类JSON字符串。
        
        Args:
            value: 要推断类型的值，可以是任何Python数据类型。
            
        Returns:
            推断出的TypeScript类型名称 ("string", "number", "boolean", "any") 或 None (如果输入是pd.NA)。
        """
        if pd.isna(value): # 首先处理Pandas的NA值
            return None
        if isinstance(value, bool): # 布尔值
            return "boolean"
        if isinstance(value, (int, float)): # 整数或浮点数
            # 不需要区分int和float，TypeScript统一为number
            return "number"
        if isinstance(value, str): # 字符串需要进一步判断
            value_stripped = value.strip() # 去除首尾空格
            if value_stripped == "": # 空字符串认为是string类型
                 return "string"

            # 尝试判断是否像数字 (允许负号、逗号、小数点、科学计数法)
            # 正则表达式解释:
            # ^-?       : 可选的负号开头
            # \d{1,3}   : 1到3位数字（千分位第一组）
            # (?:,?\d{3})*: 零或多个逗号（可选）后跟3位数字（千分位后续组）
            # (?:\.\d+)? : 可选的小数点后跟至少一位数字
            # (?:[eE][+-]?\d+)?$: 可选的科学计数法部分结尾
            if re.fullmatch(r'-?\d{1,3}(?:,?\d{3})*(?:\.\d+)?(?:[eE][+-]?\d+)?', value_stripped):
                 # 正则匹配成功，尝试移除逗号并转换为float，确保不是 '1,2,3' 这种无效数字
                 try:
                      float(value_stripped.replace(',', ''))
                      return "number" # 转换成功，是数字
                 except ValueError:
                      # 转换失败，说明虽然正则匹配，但格式无效（如多个小数点，逗号位置不对等）
                      # 此时应视为普通字符串
                      pass # 继续执行后续判断

            # 判断是否是布尔字符串 "true" 或 "false" (不区分大小写)
            if value_stripped.lower() in ('true', 'false'):
                return "boolean"

            # 判断是否像JSON对象或数组字符串
            if (value_stripped.startswith('{') and value_stripped.endswith('}')) or \
               (value_stripped.startswith('[') and value_stripped.endswith(']')):
                # 尝试使用更健壮的方式解析类JSON字符串
                try:
                    # 1. 给没有引号的键添加双引号: key: -> "key":
                    processed_value = re.sub(r'(?<!")\b(\w+)\b(?=\s*:)', r'"\1"', value_stripped)
                    # 2. 处理单引号为双引号 (如果需要)
                    # processed_value = processed_value.replace("'", '"')
                    # 3. 处理末尾多余的逗号: ,} -> } 或 ,] -> ]
                    processed_value = re.sub(r',\s*([}\]])', r'\1', processed_value)
                    # 4. 尝试用 literal_eval 安全解析
                    literal_eval(processed_value)
                    # 解析成功，说明是有效的对象或数组结构，归为 "any" 类型
                    return "any"
                except (SyntaxError, ValueError, TypeError) as e:
                    # 解析失败，说明它只是长得像JSON的普通字符串
                    # self.logger.debug(f"String '{value}' resembles JSON but failed to parse: {e}. Treating as string.")
                    pass # 继续执行，当作普通字符串处理

            # 如果以上都不是，则认为是普通字符串
            return "string"

        # 对于其他Python内置类型（如datetime, bytes等）或自定义对象
        # 无法直接映射到基本的TS类型，归为 "any"
        # self.logger.debug(f"Value '{value}' of type {type(value)} cannot be directly mapped, inferring as 'any'.")
        return "any"

    def _parse_data_rows(self, df: pd.DataFrame,
                         metadata: Dict[str, Any]) -> List[Dict]:
        """
        解析数据行（从第三行开始），将DataFrame中的数据转换为字典列表。
        使用metadata中的列信息来确定键和处理值。
        
        Args:
            df: DataFrame数据，包含所有行。
            metadata: 元数据字典，必须包含 "columns": List[ColumnInfo] 和内部使用的 "_key_map"。
            
        Returns:
            数据行列表，每个元素是一个字典，表示一行有效数据。
        """
        data_rows = [] # 存储最终的行数据字典列表
        # 获取列信息列表和内部使用的 key 到 ColumnInfo 的映射
        columns_info: List[ColumnInfo] = metadata.get("columns", [])
        key_map = metadata.get("_key_map", {}) # 主要用于查找原始列索引

        # 如果没有有效的列信息，无法解析数据行
        if not columns_info:
             self.logger.warning(f"在工作表 '{metadata.get('sheet_name', 'Unknown')}' 中没有解析到有效的列信息（键名），无法处理数据行。")
             return []

        # 如果DataFrame行数不足3，说明没有数据行
        if len(df) < 3:
            return []

        # 创建一个映射：ColumnInfo对象 -> 它在原始DataFrame中的列索引(j)
        # 这有助于直接通过ColumnInfo访问对应的原始列数据，避免每次都查找
        col_to_df_index = {}
        raw_keys_list = df.iloc[1].tolist() # 获取第二行的键名列表
        for col_info in columns_info:
            try:
                # 找到第一个匹配该键名的列索引
                df_index = next(j for j, k in enumerate(raw_keys_list) if pd.notna(k) and str(k).strip() == col_info.key)
                col_to_df_index[col_info] = df_index
            except StopIteration:
                # 如果在原始键名行中找不到该键（理论上不应发生，因为columns_info是从那里生成的）
                self.logger.error(f"严重错误：无法在原始键名行中找到键 '{col_info.key}' 对应的列索引。该列数据将被忽略。")
                col_to_df_index[col_info] = -1 # 标记为无效索引


        # 遍历DataFrame中的数据行，从第三行开始（索引2）
        # df.iloc[2:] 选择从第三行到末尾的所有行
        for idx, row_series in df.iloc[2:].iterrows():
            # idx 是原始DataFrame的行索引 (从0开始)
            # row_series 是一个Pandas Series，表示当前行的数据

            # 检查是否为注释行（根据第一列的值判断）
            # 需要确保行中有数据才访问iloc[0]
            first_cell_value = row_series.iloc[0] if len(row_series) > 0 else None
            # 如果第一列为空值 (NaN) 或者 是以 "//" 开头的字符串，则跳过该行
            if pd.isna(first_cell_value) or \
               (isinstance(first_cell_value, str) and first_cell_value.strip().startswith("//")):
                continue # 跳到下一行

            row_data = {} # 用于存储当前行数据的字典 {key: processed_value}
            # 遍历元数据中定义的每一列信息
            for col_info in columns_info:
                key = col_info.key         # 当前列的键名
                target_type = col_info.type # 当前列的目标类型

                # 获取该列在原始DataFrame中的索引
                original_col_index = col_to_df_index.get(col_info, -1)

                # 如果索引无效或超出当前行的范围，则跳过该列
                if original_col_index == -1 or original_col_index >= len(row_series):
                    continue

                # 获取原始单元格的值
                raw_value = row_series.iloc[original_col_index]

                # 跳过空值单元格(NaN)
                if pd.isna(raw_value):
                    # 可以选择是否在输出中包含null或省略该键
                    # row_data[key] = None # 如果需要显式表示null
                    continue # 当前选择省略空值

                # 使用目标类型处理原始值
                processed_value = self._value_handle(raw_value, target_type)

                # 只有在值处理后有意义时才添加到行数据字典中
                # _value_handle 对于无法处理的值可能返回 None
                if processed_value is not None:
                    row_data[key] = processed_value
                # else:
                    # 可以选择记录处理失败的情况
                    # self.logger.debug(f"Value '{raw_value}' in row {idx+1}, key '{key}' resulted in None after handling for type '{target_type}'.")


            # 只有当行数据字典不为空时，才将其添加到结果列表中
            # 避免添加完全由空值或注释组成的行（虽然注释行已在前面跳过）
            if row_data:
                 data_rows.append(row_data)

        return data_rows # 返回包含所有有效数据行字典的列表

    def _value_handle(self, value: Any, target_type: str) -> Any:
        """
        处理从Excel单元格读取的原始值，根据目标TypeScript类型进行转换和格式化。
        
        Args:
            value: 单元格的原始值 (保证调用时不是 pd.NA)。
            target_type: 从ColumnInfo获取的目标TypeScript类型 ("string", "number", "boolean", "any")。
            
        Returns:
            处理后的值，适用于最终生成的JSON/TypeScript数据结构，或者None（如果无法处理或应被视为空）。
        """
        # 根据目标类型进行处理
        if target_type == "number":
            if isinstance(value, (int, float)):
                return value # 已经是数字，直接返回
            if isinstance(value, str):
                # 尝试将字符串转换为数字 (移除逗号)
                value_cleaned = value.strip().replace(',', '')
                try:
                    float_val = float(value_cleaned)
                    # 如果是整数形式的浮点数（如 5.0），可以转换为整数
                    return int(float_val) if float_val.is_integer() else float_val
                except ValueError:
                    # 转换失败，记录警告并可能返回None或原始值
                    self.logger.warning(f"值 '{value}' 期望为 number 类型，但无法从字符串转换。将返回 None。")
                    return None # 表示转换失败
            # 其他类型（如bool）通常不应强制转为number
            self.logger.warning(f"值 '{value}' (类型 {type(value)}) 期望为 number 类型，但无法转换。将返回 None。")
            return None

        elif target_type == "boolean":
            if isinstance(value, bool):
                return value # 已经是布尔值
            if isinstance(value, (int, float)):
                 # Python中，0和0.0为False，其他数字为True
                 return bool(value)
            if isinstance(value, str):
                 # 处理字符串形式的布尔值
                 value_lower = value.strip().lower()
                 if value_lower == 'true':
                      return True
                 if value_lower == 'false':
                      return False
                 # 也可以考虑将 '1' 转 True, '0' 转 False，但需谨慎
                 # if value_lower == '1': return True
                 # if value_lower == '0': return False
            # 其他类型无法可靠转为 boolean
            self.logger.warning(f"值 '{value}' (类型 {type(value)}) 期望为 boolean 类型，但无法可靠转换。将返回 None。")
            return None

        elif target_type == "string":
             # 目标是字符串，几乎所有类型都转为字符串表示
             if isinstance(value, str):
                  # 对字符串进行清理，去除首尾空格即可
                  return value.strip()
             # 对于数字、布尔等，直接使用str()转换
             return str(value)

        elif target_type == "any":
            # 对于 'any' 类型，尝试保留原始结构，特别是解析类JSON字符串
            if isinstance(value, str):
                value_stripped = value.strip()
                # 检查是否像对象或数组
                if (value_stripped.startswith('{') and value_stripped.endswith('}')) or \
                   (value_stripped.startswith('[') and value_stripped.endswith(']')):
                    try:
                        # 尝试解析为Python对象 (字典或列表)
                        # 与 infer_type_single 中类似的解析逻辑
                        processed_value = re.sub(r'(?<!")\b(\w+)\b(?=\s*:)', r'"\1"', value_stripped)
                        processed_value = re.sub(r',\s*([}\]])', r'\1', processed_value)
                        parsed_obj = literal_eval(processed_value)
                        # 解析成功，返回Python对象，后续会被json库正确序列化
                        return parsed_obj
                    except (SyntaxError, ValueError, TypeError) as e:
                         # 解析失败，记录警告，并将其作为普通字符串返回
                         self.logger.warning(f"尝试将目标类型为 'any' 的字符串 '{value}' 解析为对象/数组失败: {e}. 将其作为普通字符串处理。")
                         return value_stripped # 返回清理后的原始字符串
                else:
                     # 如果不是明显的 JSON 结构，就当作普通字符串返回
                     return value_stripped
            # 对于已经是数字、布尔值或其他非字符串类型的，直接返回值
            # 这些类型通常能被JSON库直接处理
            return value

        # 如果 target_type 不是预期的类型（代码逻辑问题）
        self.logger.error(f"内部错误：未知的目标类型 '{target_type}' 用于处理值 '{value}'。")
        return value # 作为后备，返回原始值

    def _extract_colors(self, file_path: str,
                        sheet_name: str) -> Dict[str, str]:
        """
        提取Excel文件中指定工作表的单元格背景颜色信息。
        现在返回RGB十六进制字符串 (例如 'FF0000' 代表红色)。
        
        Args:
            file_path: Excel文件路径。
            sheet_name: 工作表名称。
            
        Returns:
            一个字典，键为单元格坐标字符串（如 'A1'），值为单元格背景色的RGB十六进制字符串（如 'FF0000'）。
            如果单元格无填充或提取失败，则不包含该单元格。
        """
        try:
            # 加载工作簿，data_only=False 以便访问样式信息
            workbook = openpyxl.load_workbook(file_path, data_only=False, read_only=True) # read_only=True 可能提高性能
            if sheet_name not in workbook.sheetnames:
                 self.logger.warning(f"在文件 '{os.path.basename(file_path)}' 中找不到工作表 '{sheet_name}'，无法提取颜色。")
                 return {}
            sheet = workbook[sheet_name]

            color_map = {} # 存储单元格坐标到颜色RGB值的映射
            # 遍历工作表中的所有单元格
            # sheet.max_row 和 sheet.max_column 可能不准确，iter_rows更可靠
            for row in sheet.iter_rows(): # 迭代所有行
                for cell in row: # 迭代行中的所有单元格
                    # 检查单元格是否有填充 (fill) 并且填充类型是 'solid' (纯色)
                    # 并且有关联的前景色 (fgColor)
                    if cell.fill and cell.fill.fill_type == 'solid' and cell.fill.fgColor:
                         color_info = cell.fill.fgColor # 获取前景色对象

                         # 颜色的类型可能是 'rgb', 'indexed', 'theme'
                         # 我们主要关心 'rgb'
                         if color_info.type == 'rgb':
                              # 获取ARGB值 (Alpha, Red, Green, Blue)，格式通常是 AARRGGBB
                              argb_color = color_info.rgb
                              # 验证颜色值有效且不是默认的透明或无填充 (通常是 '00000000')
                              if argb_color and len(argb_color) == 8 and argb_color != '00000000':
                                   # 我们通常只需要RGB部分，即去掉前两位Alpha值
                                   rgb_color = argb_color[2:]
                                   # 存储坐标和RGB颜色值
                                   color_map[cell.coordinate] = rgb_color # 例如 'FF0000'
                         # 可以选择性处理其他颜色类型，但通常比较复杂或不准确
                         # elif color_info.type == 'indexed':
                         #     # 索引颜色需要映射表，openpyxl可能不直接提供准确RGB
                         #     self.logger.debug(f"Indexed color {color_info.value} at {cell.coordinate}. Extraction might be inaccurate.")
                         # elif color_info.type == 'theme':
                         #     # 主题颜色需要解析文档主题部分
                         #     self.logger.debug(f"Theme color at {cell.coordinate}. Extraction not fully supported.")

            # 关闭工作簿文件句柄（如果以read_only=False打开）
            # workbook.close() # 在 read_only=True 模式下不需要 close
            return color_map
        except Exception as e:
            # 记录提取颜色时发生的错误
            self.logger.error(f"提取单元格颜色时出错: {e}", exc_info=True)
            # 包含文件和工作表信息以帮助调试
            self.logger.error(f"File path: {file_path}, Sheet name: {sheet_name}")
            return {} # 出错时返回空字典
