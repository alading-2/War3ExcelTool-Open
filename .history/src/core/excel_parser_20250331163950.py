import os  # 导入操作系统模块，用于处理文件路径
import pandas as pd  # 导入pandas库，简称pd，用于数据分析和处理Excel文件
import openpyxl  # 导入openpyxl库，用于读取Excel文件的样式信息（如颜色）
from typing import Dict, List, Any, Tuple, Optional, Union  # 导入类型提示功能
from openpyxl.styles import PatternFill  # 导入Excel单元格填充样式相关功能
import logging  # 导入日志模块
from src.utils.logger import get_logger  # 从项目的utils目录导入自定义日志获取函数

# 获取当前模块的日志记录器
logger = get_logger(__name__)  # __name__是Python的特殊变量，表示当前模块名称


class ExcelParser:
    """解析Excel文件并提取数据的类
    
    这个类负责读取Excel文件的内容，提取其中的数据、结构和颜色信息，
    为后续生成TypeScript代码做准备。
    """

    def __init__(self):
        """初始化ExcelParser类的新实例
        
        创建一个日志记录器，用于记录Excel解析过程中的信息
        """
        # 获取一个与当前类关联的日志记录器
        self.logger = logging.getLogger(__name__)

    def parse_excel(
        self, file_path: str
    ) -> List[Tuple[Dict[str, Any], List[str], List[Dict], Dict[str, str]]]:
        """
        解析Excel文件，提取数据和结构信息
        
        这个方法是整个解析过程的主入口，它会读取Excel文件并提取多种信息：
        1. 元数据（如注释、键名等）
        2. 数据字段的键名列表
        3. 实际数据行的列表
        4. 单元格颜色信息
        
        Args:
            file_path: Excel文件路径，表示要解析的Excel文件的位置
            sheet_name: 可选的工作表名称，如果不提供则解析所有工作表
            
        Returns:
            如果指定了sheet_name，返回一个包含四个元素的元组(Tuple)：
            - 元数据: 字典，包含注释、键名列表和文件名等信息
            - 键列表: 列表，数据字段的键名列表（不包括ID列）
            - 数据列表: 列表，每个元素是一个字典，表示一行数据
            - 颜色映射: 字典，单元格位置到颜色代码的映射
            
            如果没有指定sheet_name，返回一个列表，每个元素包含五个元素的元组(Tuple)：
            - sheet名称: 字符串，工作表的名称
            - 元数据: 字典，包含注释、键名列表和文件名等信息
            - 键列表: 列表，数据字段的键名列表（不包括ID列）
            - 数据列表: 列表，每个元素是一个字典，表示一行数据
            - 颜色映射: 字典，单元格位置到颜色代码的映射
            
        Raises:
            FileNotFoundError: 当指定的Excel文件不存在时
            Exception: 当解析过程中发生其他错误时
        """
        # 检查文件是否存在，不存在则记录错误并抛出异常
        if not os.path.exists(file_path):
            self.logger.error(f"Excel文件不存在: {file_path}")
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")

        try:
            # 获取Excel文件的所有sheet名称
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names

            results = []

            # 遍历所有sheet并解析
            for sheet_name in sheet_names:
                self.logger.info(f"解析工作表: {sheet_name}")

                # 使用pandas读取指定sheet的内容
                df = pd.read_excel(file_path,
                                   sheet_name=sheet_name,
                                   header=None)

                # 提取第一行作为注释信息（列说明）
                comments = df.iloc[0].tolist()

                # 提取第二行作为数据的键名
                keys = df.iloc[1].tolist()

                # 解析实际数据行（第三行及以下）
                data_rows = self._parse_data_rows(df, keys)

                # 读取当前sheet的单元格颜色信息
                color_map = self._extract_colors(file_path, sheet_name)

                # 创建元数据字典
                metadata = {
                    "comments": comments,
                    "keys": keys,
                    "filename": os.path.basename(file_path),
                    "sheet_name": sheet_name  # 添加sheet名称到元数据
                }

                # 将当前sheet的解析结果添加到列表
                results.append((metadata, keys[1:], data_rows, color_map))

            return results

        except Exception as e:
            self.logger.error(f"解析Excel文件的工作表时出错: {str(e)}")
            raise

    def _parse_data_rows(self, df: pd.DataFrame, keys: List) -> List[Dict]:
        """
        解析数据行，将DataFrame中的数据转换为字典列表
        
        Args:
            df: DataFrame数据，pandas的数据表格对象
            keys: 键列表，表示每列的字段名
            
        Returns:
            数据行列表，每个元素是一个字典，表示一行数据
        """
        # 创建空列表，用于存储解析后的数据行
        data_rows = []

        # 遍历DataFrame中的行，从第三行开始（索引2）
        # idx是行索引，row是行数据
        for idx, row in df.iloc[2:].iterrows():
            # 检查是否为注释行（首列数据为空或者以"//"开头的行视为注释，将被跳过）
            # str(): 用于将对象转换为字符串。strip(): 去除字符串两端的空格
            # startswith(): 检查字符串是否以指定的前缀开头
            if row.iloc[0] == "" or str(row.iloc[0]).strip().startswith("//"):
                continue  # 跳过本次循环，处理下一行

            # 提取行ID（第一列的值）
            row_id = row.iloc[0]

            # 创建行数据字典，并设置ID字段
            row_data = {"id": row_id}

            # 遍历行中的每个单元格（从第二列开始）
            # i是列索引（从1开始），value是单元格的值
            for i, value in enumerate(row.iloc[1:], 1):
                # 跳过空数据单元格（NaN或空字符串）
                if pd.isna(value) or value == "":
                    continue

                # 检查列索引是否在键列表范围内
                if i < len(keys):
                    # 获取当前列的键名
                    key_name = keys[i]
                    # 将值添加到行数据字典中
                    row_data[key_name] = value

            # 将处理好的行数据添加到结果列表中
            data_rows.append(row_data)

        # 返回所有数据行的列表
        return data_rows

    def _extract_colors(self,
                        file_path: str,
                        sheet_name: Optional[str] = None) -> Dict[str, str]:
        """
        提取Excel中的单元格文本字体颜色信息
        
        这个方法使用openpyxl库读取Excel文件中的单元格字体颜色，
        并将其转换为|cffRRGGBB|r颜色代码格式。对于富文本单元格，
        会分析每个字符的颜色，将相同颜色的连续字符组合在一起。
        
        Args:
            file_path: Excel文件路径
            sheet_name: 可选的工作表名称，如果不提供则使用活动工作表
            
        Returns:
            颜色映射字典: {单元格位置: 带颜色标记的文本}，位置格式为"行号_列号"
        """
        try:
            # 使用openpyxl加载Excel工作簿
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # 获取指定的工作表或活动工作表
            if sheet_name and sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                sheet = wb.active

            # 创建空字典，用于存储单元格位置到颜色标记文本的映射
            color_map = {}

            # 遍历工作表中的所有单元格
            for row_idx, row in enumerate(sheet.rows, 1):
                for col_idx, cell in enumerate(row, 1):
                    # 跳过前两行（标题和键名行）
                    if row_idx <= 2:
                        continue
                        
                    # 跳过空单元格
                    if cell.value is None:
                        continue

                    cell_text = str(cell.value)

                    # 处理富文本单元格
                    if hasattr(cell, 'rich_text') and cell.rich_text:
                        formatted_text = self._process_rich_text(cell)
                        if formatted_text and formatted_text != cell_text:  # 只保存有颜色标记的文本
                            position = f"{row_idx}_{col_idx}"
                            color_map[position] = formatted_text
                    # 处理普通文本单元格
                    elif hasattr(cell, 'font') and cell.font is not None:
                        font = cell.font
                        if hasattr(font, 'color') and font.color is not None:
                            color = getattr(font.color, 'rgb', None)
                            if color:
                                # 如果颜色以'FF'开头，去除这个alpha通道前缀
                                if isinstance(color, str) and color.startswith('FF'):
                                    color = color[2:]

                                # 只有当颜色不是黑色(000000)时才添加颜色标记
                                if color.upper() != '000000':
                                    formatted_text = f"|cff{color}{cell_text}|r"
                                    position = f"{row_idx}_{col_idx}"
                                    color_map[position] = formatted_text

            # 关闭工作簿，释放资源
            wb.close()

            return color_map
        except Exception as e:
            self.logger.error(f"提取颜色信息时出错: {str(e)}")
            return {}  # 出错时返回空字典

    def _process_rich_text(self, cell) -> str:
        """
        处理富文本单元格，为每段不同颜色的文本添加颜色标记
        
        Args:
            cell: openpyxl单元格对象，包含富文本
            
        Returns:
            带颜色标记的文本字符串
        """
        try:
            if not hasattr(cell, 'rich_text') or not cell.rich_text:
                return str(cell.value)
                
            result = []
            current_color = None
            current_text = ""

            # 遍历富文本中的每个文本片段
            for fragment in cell.rich_text:
                # 检查fragment是否有text属性
                if not hasattr(fragment, 'text'):
                    continue
                    
                text = fragment.text
                if not text:
                    continue

                # 获取文本片段的颜色
                color = None
                # 安全地获取颜色属性
                if hasattr(fragment, 'font') and fragment.font is not None:
                    font = fragment.font
                    if hasattr(font, 'color') and font.color is not None:
                        color = getattr(font.color, 'rgb', None)
                        if color and isinstance(color, str) and color.startswith('FF'):
                            color = color[2:]

                # 如果颜色与当前累积文本的颜色相同，则继续累积
                if color == current_color:
                    current_text += text
                else:
                    # 如果有之前累积的文本，先处理它
                    if current_text:
                        if current_color and current_color.upper() != '000000':
                            result.append(f"|cff{current_color}{current_text}|r")
                        else:
                            result.append(current_text)

                    # 开始新的文本累积
                    current_color = color
                    current_text = text

            # 处理最后累积的文本
            if current_text:
                if current_color and current_color.upper() != '000000':
                    result.append(f"|cff{current_color}{current_text}|r")
                else:
                    result.append(current_text)

            return "".join(result)
        except Exception as e:
            self.logger.error(f"处理富文本时出错: {str(e)}")
            return str(cell.value) if hasattr(cell, 'value') else ""  # 出错时返回原始文本或空字符串

    def _character_by_character_color_processing(self, cell) -> str:
        """
        逐字符处理单元格内容，为每个字符添加颜色标记
        这是一个备用方法，当单元格不包含富文本但需要字符级颜色处理时使用
        
        Args:
            cell: openpyxl单元格对象
            
        Returns:
            带颜色标记的文本字符串
        """
        try:
            # 确保cell有value属性
            if not hasattr(cell, 'value') or cell.value is None:
                return ""
                
            # 此方法需要与Excel的COM接口配合使用，在此仅提供框架
            # 由于openpyxl不直接支持字符级颜色，此功能可能需要其他库支持
            # 例如可以使用pywin32库与Excel COM对象交互

            # 暂时简单返回单元格值，后续可以扩展
            return str(cell.value)
        except Exception as e:
            self.logger.error(f"处理字符级颜色时出错: {str(e)}")
            return str(cell.value) if hasattr(cell, 'value') else ""  # 出错时返回原始文本或空字符串
