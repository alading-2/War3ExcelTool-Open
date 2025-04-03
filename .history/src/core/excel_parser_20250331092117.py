import os  # 导入操作系统模块，用于处理文件路径
import pandas as pd  # 导入pandas库，简称pd，用于数据分析和处理Excel文件
import openpyxl  # 导入openpyxl库，用于读取Excel文件的样式信息（如颜色）
from typing import Dict, List, Any, Tuple, Optional, Union  # 导入类型提示功能
from openpyxl.styles import PatternFill  # 导入Excel单元格填充样式相关功能
import logging  # 导入日志模块
from src.utils.logger import get_logger  # 从项目的utils目录导入自定义日志获取函数

# 获取当前模块的日志记录器
logger = get_logger(__name__)  # __name__是Python的特殊变量，表示当前模块名称

class ExcelParser:
    """解析Excel文件并提取数据的类
    
    这个类负责读取Excel文件的内容，提取其中的数据、结构和颜色信息，
    为后续生成TypeScript代码做准备。
    """
    
    def __init__(self):
        """初始化ExcelParser类的新实例
        
        创建一个日志记录器，用于记录Excel解析过程中的信息
        """
        # 获取一个与当前类关联的日志记录器
        self.logger = logging.getLogger(__name__)
    
    def parse_excel(self, file_path: str)
        """
        解析Excel文件，提取数据和结构信息
        
        这个方法是整个解析过程的主入口，它会读取Excel文件并提取多种信息：
        1. 元数据（如注释、键名等）
        2. 数据字段的键名列表
        3. 实际数据行的列表
        4. 单元格颜色信息
        
        Args:
            file_path: Excel文件路径，表示要解析的Excel文件的位置
            sheet_name: 可选的工作表名称，如果不提供则解析所有工作表
            
        Returns:
            如果指定了sheet_name，返回一个包含四个元素的元组(Tuple)：
            - 元数据: 字典，包含注释、键名列表和文件名等信息
            - 键列表: 列表，数据字段的键名列表（不包括ID列）
            - 数据列表: 列表，每个元素是一个字典，表示一行数据
            - 颜色映射: 字典，单元格位置到颜色代码的映射
            
            如果没有指定sheet_name，返回一个列表，每个元素包含五个元素的元组(Tuple)：
            - sheet名称: 字符串，工作表的名称
            - 元数据: 字典，包含注释、键名列表和文件名等信息
            - 键列表: 列表，数据字段的键名列表（不包括ID列）
            - 数据列表: 列表，每个元素是一个字典，表示一行数据
            - 颜色映射: 字典，单元格位置到颜色代码的映射
            
        Raises:
            FileNotFoundError: 当指定的Excel文件不存在时
            Exception: 当解析过程中发生其他错误时
        """
        # 检查文件是否存在，不存在则记录错误并抛出异常
        if not os.path.exists(file_path):
            self.logger.error(f"Excel文件不存在: {file_path}")
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        try:
            # 获取Excel文件的所有sheet名称
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names
                        
            results = []
            
            # 遍历所有sheet并解析
            for sheet_name in sheet_names:
                self.logger.info(f"解析工作表: {sheet_name}")
                
                # 使用pandas读取指定sheet的内容
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                
                # 提取第一行作为注释信息（列说明）
                comments = df.iloc[0].tolist()
                
                # 提取第二行作为数据的键名
                keys = df.iloc[1].tolist()
                
                # 解析实际数据行（第三行及以下）
                data_rows = self._parse_data_rows(df, keys)
                
                # 读取当前sheet的单元格颜色信息
                color_map = self._extract_colors(file_path, sheet_name)
                
                # 创建元数据字典
                metadata = {
                    "comments": comments,
                    "keys": keys,
                    "filename": os.path.basename(file_path),
                    "sheet_name": sheet_name  # 添加sheet名称到元数据
                }
                
                # 将当前sheet的解析结果添加到列表
                results.append((metadata, keys[1:], data_rows, color_map))
            
            return results
            
        except Exception as e:
            self.logger.error(f"解析Excel文件的工作表时出错: {str(e)}")
            raise

    def _parse_data_rows(self, df: pd.DataFrame, keys: List) -> List[Dict]:
        """
        解析数据行，将DataFrame中的数据转换为字典列表
        
        Args:
            df: DataFrame数据，pandas的数据表格对象
            keys: 键列表，表示每列的字段名
            
        Returns:
            数据行列表，每个元素是一个字典，表示一行数据
        """
        # 创建空列表，用于存储解析后的数据行
        data_rows = []
        
        # 遍历DataFrame中的行，从第三行开始（索引2）
        # idx是行索引，row是行数据
        for idx, row in df.iloc[2:].iterrows():
            # 检查是否为注释行（首列以"//"开头的行视为注释，将被跳过）
            if str(row.iloc[0]).strip().startswith("//"):
                continue  # 跳过本次循环，处理下一行
                
            # 提取行ID（第一列的值）
            row_id = row.iloc[0]  
            
            # 创建行数据字典，并设置ID字段
            row_data = {"id": row_id}
            
            # 遍历行中的每个单元格（从第二列开始）
            # i是列索引（从1开始），value是单元格的值
            for i, value in enumerate(row.iloc[1:], 1):
                # 跳过空数据单元格（NaN或空字符串）
                if pd.isna(value) or value == "":
                    continue
                    
                # 检查列索引是否在键列表范围内
                if i < len(keys):
                    # 获取当前列的键名
                    key_name = keys[i]
                    # 将值添加到行数据字典中
                    row_data[key_name] = value
                    
            # 将处理好的行数据添加到结果列表中
            data_rows.append(row_data)
        
        # 返回所有数据行的列表
        return data_rows
    
    def _extract_colors(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, str]:
        """
        提取Excel中的单元格颜色信息
        
        这个方法使用openpyxl库读取Excel文件中的单元格填充颜色，
        并将其转换为TypeScript中使用的颜色代码格式。
        
        Args:
            file_path: Excel文件路径
            sheet_name: 可选的工作表名称，如果不提供则使用活动工作表
            
        Returns:
            颜色映射字典: {单元格位置: 颜色代码}，位置格式为"行号_列号"
        """
        
        # 使用openpyxl加载Excel工作簿，data_only=True表示读取值而非公式
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 获取指定的工作表或活动工作表
        sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        
        # 创建空字典，用于存储单元格位置到颜色代码的映射
        color_map = {}
        
        # 遍历工作表中的行（从第三行开始，即实际数据行）
        # row_idx是行索引（从3开始），row是行对象
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3), 3):
            # 遍历行中的每个单元格
            # col_idx是列索引（从1开始），cell是单元格对象
            for col_idx, cell in enumerate(row, 1):
                # 只关注有填充颜色的单元格，使用命名表达式优化
                # 条件1: 单元格有实色填充
                # 条件2: 填充颜色不是默认的无色
                # 条件3: 成功提取到颜色RGB值（使用:=赋值表达式）
                if (cell.fill.fill_type == "solid" and 
                    cell.fill.start_color.index != '00000000' and 
                    (color := cell.fill.start_color.rgb)):  # := 是Python 3.8引入的赋值表达式，在条件中赋值并使用变量
                    
                    # 转换为我们需要的颜色格式
                    # 如果颜色以'FF'开头，去除这个alpha通道前缀
                    if color.startswith('FF'):
                        color = color[2:]
                        
                    # 构造单元格位置标识符（格式：行号_列号）
                    position = f"{row_idx}_{col_idx}"
                    
                    # 将位置和颜色代码添加到颜色映射字典中
                    color_map[position] = f"|cff{color}"
        
        # 关闭工作簿，释放资源
        wb.close()
        
        # 返回颜色映射字典
        return color_map
