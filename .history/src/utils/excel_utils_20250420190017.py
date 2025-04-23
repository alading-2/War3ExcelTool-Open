import os
import win32com.client
import pythoncom
import time
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging

def ensure_excel_file_closed(file_path):
    """
    确保Excel文件未被打开，如果已打开则强制关闭
    
    Args:
        file_path: Excel文件路径
    """
    # 标准化文件路径
    file_path = os.path.abspath(file_path)
    
    # 如果文件不存在，不需要进一步处理
    if not os.path.exists(file_path):
        return
    
    try:
        # 尝试以写模式打开文件，检查文件是否已被锁定
        with open(file_path, 'r+b'):
            # 文件未被锁定，可以自由访问
            pass
    except IOError:
        # 文件被锁定，尝试通过Excel应用程序关闭它
        try:
            # 初始化COM环境
            pythoncom.CoInitialize()
            
            # 获取Excel应用程序实例
            excel = win32com.client.Dispatch("Excel.Application")
            
            # 检查所有已打开的工作簿
            for wb in excel.Workbooks:
                # 标准化工作簿的完整路径
                wb_path = os.path.abspath(wb.FullName)
                
                # 如果找到匹配的工作簿，关闭它(不保存更改)
                if wb_path.lower() == file_path.lower():
                    wb.Close(SaveChanges=False)
                    print(f"已关闭打开的Excel文件: {file_path}")
            
            # 如果没有其他打开的工作簿，退出Excel应用程序
            if excel.Workbooks.Count == 0:
                excel.Quit()
            
            # 释放COM对象
            excel = None
            
            # 短暂等待以确保文件被完全释放
            time.sleep(0.5)
            
        except Exception as e:
            print(f"尝试关闭Excel文件时出错: {e}")
        finally:
            # 释放COM环境
            pythoncom.CoUninitialize()

def format_excel_workbook(file_path):
    """
    优化Excel工作簿中所有工作表的格式
    
    应用以下格式:
    1. 添加所有框线
    2. 字体大小：14
    3. 水平居中垂直居中
    4. sheet首行背景颜色改为浅绿色
    5. 禁止文本自动换行
    6. 禁止自动缩小字体适应单元格
    7. 冻结窗格到1行A列
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        bool: 格式化成功返回True，失败返回False
    """
    logger = logging.getLogger(__name__)
    
    try:
        # 确保文件未被打开
        ensure_excel_file_closed(file_path)
        
        # 加载工作簿
        workbook = load_workbook(file_path)
        
        # 定义样式
        # 创建边框样式 - 设置单元格四周边框为细线
        thin_border = Border(
            left=Side(style='thin'),    # 左边框为细线
            right=Side(style='thin'),   # 右边框为细线
            top=Side(style='thin'),     # 上边框为细线
            bottom=Side(style='thin')   # 下边框为细线
        )
        
        # 创建字体样式 - 设置字体大小为14
        font = Font(size=14)  # 字号14，便于阅读
        
        # 创建对齐方式 - 设置单元格内容水平垂直居中且不自动换行
        alignment = Alignment(
            horizontal='center',    # 水平居中对齐
            vertical='center',      # 垂直居中对齐
            wrap_text=False,        # 禁止文本自动换行
            shrink_to_fit=False     # 禁止自动缩小字体以适应单元格
        )
        
        # 创建首行填充样式 - 设置浅绿色背景突出表头
        header_fill = PatternFill(
            start_color='C6EFCE',   # 浅绿色的十六进制颜色代码
            end_color='C6EFCE',     # 确保颜色一致
            fill_type='solid'       # 实心填充样式
        )
        
        # 处理每个工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 获取表格的尺寸
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # 调整列宽以适应内容
            for col in range(1, max_col + 1):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = 20  # 设置默认宽度
            
            # 应用单元格格式
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    
                    # 应用边框
                    cell.border = thin_border
                    
                    # 应用字体
                    cell.font = font
                    
                    # 应用对齐方式
                    cell.alignment = alignment
                    
                    # 对首行应用填充色
                    if row == 1:
                        cell.fill = header_fill
            
            # 冻结窗格到1行A列（B2单元格的左侧和上方）
            sheet.freeze_panes = 'B2'
        
        # 保存修改
        workbook.save(file_path)
        logger.info(f"成功优化Excel文件格式: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"格式化Excel文件失败: {str(e)}")
        return False 