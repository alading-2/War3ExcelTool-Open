import os
import pandas as pd
import openpyxl
from typing import Dict, List, Any, Tuple, Optional
from openpyxl.styles import PatternFill
import logging
from ..utils.logger import get_logger

logger = get_logger(__name__)

class ExcelParser:
    """解析Excel文件并提取数据的类"""
    
    def __init__(self):
        """初始化ExcelParser类"""
        self.logger = logging.getLogger(__name__)
    
    def parse_excel(self, file_path: str) -> Tuple[Dict[str, Any], List[str], List[Dict], Dict[str, str]]:
        """
        解析Excel文件，提取数据和结构信息
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Tuple包含:
            - 元数据: 包含注释、键等信息
            - 键列表: 数据键名列表
            - 数据列表: 实际数据对象列表
            - 颜色映射: 单元格颜色信息
        """
        if not os.path.exists(file_path):
            self.logger.error(f"Excel文件不存在: {file_path}")
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        try:
            # 读取数据内容
            df = pd.read_excel(file_path, header=None)
            
            # 读取颜色信息
            color_map = self._extract_colors(file_path)
            
            # 第一行作为注释
            comments = df.iloc[0].tolist()
            
            # 第二行作为数据的key值
            keys = df.iloc[1].tolist()
            
            # 实际数据（第三行及以下）
            data_rows = self._parse_data_rows(df, keys)
            
            metadata = {
                "comments": comments,
                "keys": keys,
                "filename": os.path.basename(file_path)
            }
            
            return metadata, keys[1:], data_rows, color_map
            
        except Exception as e:
            self.logger.error(f"解析Excel文件时出错: {str(e)}")
            raise
    
    def _parse_data_rows(self, df: pd.DataFrame, keys: List) -> List[Dict]:
        """
        解析数据行
        
        Args:
            df: DataFrame数据
            keys: 键列表
            
        Returns:
            数据行列表
        """
        data_rows = []
        for idx, row in df.iloc[2:].iterrows():
            # 检查是否为注释行 (首列为 "//" 的行)
            if str(row.iloc[0]).strip().startswith("//"):
                continue
                
            row_id = row.iloc[0]  # 第一列作为ID
            row_data = {"id": row_id}
            
            for i, value in enumerate(row.iloc[1:], 1):
                # 跳过空数据单元格
                if pd.isna(value) or value == "":
                    continue
                    
                if i < len(keys):
                    key_name = keys[i]
                    row_data[key_name] = value
                    
            data_rows.append(row_data)
        
        return data_rows
    
    def _extract_colors(self, file_path: str) -> Dict[str, str]:
        """
        提取Excel中的单元格颜色信息
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            颜色映射字典: {单元格位置: 颜色代码}
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        color_map = {}
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3), 3):  # 从第三行开始，实际数据行
            for col_idx, cell in enumerate(row, 1):
                # 只关注有填充颜色的单元格，使用命名表达式优化
                if (cell.fill.fill_type == "solid" and 
                    cell.fill.start_color.index != '00000000' and 
                    (color := cell.fill.start_color.rgb)):
                    
                    # 转换为我们需要的颜色格式
                    if color.startswith('FF'):  # 去除alpha通道前缀
                        color = color[2:]
                    position = f"{row_idx}_{col_idx}"
                    color_map[position] = f"|cff{color}"
        
        wb.close()
        return color_map 